# coding: utf8

import os
import json
from openpyxl import Workbook


class ParselJSON:
    """
    Преобразует файл json в словарь для структурированной таблицы
    """
    def __init__(self, file):
        self.json_file = file

    def parse(self) -> dict:
        """
        :return: Словарь, где
        ключ - заголовки таблицы,
        значение - список, представляющий сзначения - строки для заголовка таблицы
        """
        headers = {}
        values = {}
        result = {}
        try:
            for header in self.json_file['headers']:
                headers[header['properties']['X']] = header['properties']['QuickInfo']
            for value in self.json_file['values']:
                if value['properties']['X'] not in values:
                    values[value['properties']['X']] = {}
                values[value['properties']['X']][value['properties']['Y']] = value['properties']['Text']
            # Совмещаем заголовки и значения в сортированном порядке
            for key in headers:
                result[headers[key]] = []
                # Получаем словарь, состоящий из кортежа, где первый элемент - порядок, второй - значение
                value = sorted(values[key].items())
                # В цикле для "универсальности", если не две строки например, а больше
                for row in value:
                    result[headers[key]].append(row[1])
        except KeyError as e:
            print(f'[ERROR]: json файл неверный. Нет ключевого слова {e}')
        return result


class XSLSWriter:
    """
    Класс записи данных в эксель
    """
    def __init__(self):
        """
        Инициализируем книгу эксель для записи
        """
        self.book = Workbook()
        # Акстивируем страницы и стандартную страницу
        ws = self.book.active
        self.book.remove(ws)

    def write_to_sheet(self, name: str, data: dict):
        """
        Создание листа эксель и запись данных
        :param name: Имя листа
        :param data: Данные для записи
        :return: None
        """
        sheet = self.book.create_sheet(name)
        sheet.title = name
        # Записали заголовки
        sheet.append(list(data.keys()))
        # Записываем значения словаря в ячейки
        col = 1
        for values in data.values():
            # каждое values из словаря - строки конкретного столбца
            # словарь поступает уже в правильном порядке
            row = 2
            for value in values:
                sheet.cell(row=row, column=col, value=value)
                row += 1
            col += 1

    def save(self, book_name):
        self.book.save(book_name)
        self.book.close()


if __name__ == '__main__':
    json_files = [file_name for file_name in os.listdir() if file_name.endswith('.json')]
    if json_files:
        writer = XSLSWriter()
        for file in json_files:
            with open(file, encoding='utf8') as f:
                json_data = json.load(f)
            parse_data = ParselJSON(json_data).parse()
            if parse_data:
                name = file.split('.')[0]
                writer.write_to_sheet(name, parse_data)
        if writer.book.worksheets:
            writer.save('book.xlsx')
    else:
        print('[ERROR]: Нет подходящих файлов для парсинга.')
